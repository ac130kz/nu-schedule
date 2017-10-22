#!/usr/bin/python3
# -*- coding: utf-8 -*-

# TODO: linking!
# improve database reading (watch INFO below)

from collections import defaultdict
from itertools import groupby, product, chain, combinations
from logging import basicConfig, DEBUG, info
from re import compile as compiles
from sys import exit, argv
from time import strptime

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from PyQt5.QtGui import QIcon
from PyQt5.QtWidgets import (QApplication, QComboBox, QDesktopWidget,
                             QDialogButtonBox, QFileDialog, QHBoxLayout, QGridLayout,
                             QLabel, QPushButton, QVBoxLayout, QWidget, QDialog, QFrame)
from PyQt5.QtCore import Qt

# Simple logging suite
basicConfig(
    filename='main.log',
     level=DEBUG,
     format='%(asctime)s - %(levelname)s - %(message)s',
     datefmt='%d/%m/%Y %I:%M:%S %p')

# Regex expression to manage courses s/t
reg = compiles('(?:\d+)([a-zA-Z]+)')

class Course():

    """Models course class"""

    def __init__(self, abbr, st, title, credit, days, timing, teacher, room):
        self.abbr = str(abbr)
        self.st = str(st)
        self.title = str(title)
        self.credit = str(credit)
        self.days = str(days)
        self.timing = str(timing)
        self.teacher = str(teacher)
        self.room = str(room)

        a, b = self.timing.split('-')
        self.start, self.end = strptime(a, "%I:%M %p"), strptime(b, "%I:%M %p")
        table = {'M':0, 'T':1, 'W':2, 'R':3, 'F':4, 'S':5}
        self.dayslist = [table.get(i) for i in table if i in self.days]

    def __repr__(self):
        return str(self.abbr + ' | ' + self.st + ' | ' + self.title +
                   ' | ' + self.credit + ' | ' + self.days + ' | ' + self.timing + ' | ' +
                   self.teacher + ' | ' + self.room) 

    def __and__(self, other):
        return self.start <= other.end and self.end >= other.start and not set.intersection(set(self.dayslist), set(other.dayslist))

class UI(QWidget):

    """Simple Qt UI"""
    
    coursesconnector = list()
    finallist = list()

    def __init__(self, parent=None):
        super(UI, self).__init__(parent)
        self.initui()

    def center(self):
        qr = self.frameGeometry()
        cp = QDesktopWidget().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())

    def on_about_clicked(self):
        d = QDialog()
        l1 = QLabel("nu-schedule\nCourses schedule generator for NU\nApache 2.0 License\n© Mikhail Krassavin, 2017")
        b1 = QPushButton('Ok', d)
        vbox = QVBoxLayout()
        vbox.addWidget(l1)
        hbox = QHBoxLayout()
        hbox.addStretch()
        hbox.addWidget(b1)
        hbox.addStretch()
        vbox.addItem(hbox)
        d.setWindowIcon(QIcon('res/logo.ico'))
        d.setWindowTitle("About")
        d.setLayout(vbox)
        b1.clicked.connect(d.accept)
        d.exec_()

    def on_help_clicked(self):
        d = QDialog()
        l1 = QLabel("1. Get a clearly formatted course list\nfor this I recommend PDF2XL,\nhave a look at a sample list in the /samples.\n2. Select the xlsx file with the Open button, wait for it\nto load, the app can be unresponsive.\n3. With Edit button select the needed courses,\nthey will appear on the Main window.\n4. Use Generate button to generate and\n save your schedule as result.txt")
        b1 = QPushButton("Ok", d)
        vbox = QVBoxLayout()
        vbox.addWidget(l1)
        hbox = QHBoxLayout()
        hbox.addStretch()
        hbox.addWidget(b1)
        hbox.addStretch()
        vbox.addItem(hbox)
        d.setWindowIcon(QIcon('res/logo.ico'))
        d.setWindowTitle("Help")
        d.setLayout(vbox)
        b1.clicked.connect(d.accept)
        d.exec_()

    def on_add_clicked(self):
        if self.coursesconnector:
            d = QDialog()
            b1 = QPushButton("Add", d)
            b2 = QPushButton("Delete", d)
            cmb1 = QComboBox()
            cmb1.addItems(sorted([t[0] for t in self.coursesconnector]))
            hbox = QHBoxLayout()
            hbox.addWidget(cmb1)
            hbox.addStretch()
            hbox.addWidget(b1)
            hbox.addWidget(b2)
            hbox.addStretch()
            d.setWindowIcon(QIcon('res/logo.ico'))
            d.setWindowTitle("Adding courses")
            d.setLayout(hbox)
            b1.clicked.connect(lambda: self.on_b1_clicked(cmb1.currentText()))
            b2.clicked.connect(lambda: self.on_b2_clicked(cmb1.currentText()))
            d.exec_()

    def done_dialog(self):
        d = QDialog()
        b1 = QPushButton("Ok", d)
        lbl1 = QLabel("Successfully loaded the database!")
        vbox = QVBoxLayout()
        vbox.addWidget(lbl1)
        vbox.addStretch()
        vbox.addWidget(b1)
        vbox.addStretch()
        d.setWindowTitle("Success!")
        d.setLayout(vbox)
        b1.clicked.connect(d.accept)
        d.setWindowIcon(QIcon('res/logo.ico'))
        d.exec_()

    def on_b1_clicked(self, text):
        if not any(text == x[0].abbr for x in self.finallist):
            for k, v in self.coursesconnector:
                if k == text:
                    for i in v:
                        self.finallist.append(i)
            
            self.label.setText('Current list is ' + str(set(y[0].abbr for y in self.finallist)))

    def on_b2_clicked(self, text):
        for x in self.finallist:
            if text == x[0].abbr:
                self.finallist.remove(x)
        self.label.setText('Current list is ' + str(set(y[0].abbr for y in self.finallist)))

    def on_open_clicked(self):
        try:
            self.label.setText('Loading... Please wait, the app can be unresponsive')
            name = QFileDialog.getOpenFileName(self, 'Open File')
            book = load_workbook(str(name[0]), data_only=True, read_only=True)
            sheet = book.worksheets[0]
            if self.checkxl(sheet):
                info(
                    'Input file ' + str(name[0]) + ' was successfully read.')

                courses = list()
                
                for i in range(2, sheet.max_row):
                    courses.append(
                        Course(
                            sheet.cell(row=i, column=1).value, sheet.cell(
                                row=i, column=2).value,
                            sheet.cell(row=i, column=3).value, sheet.cell(
                                row=i, column=5).value, sheet.cell(
                                row=i, column=8).value,
                            sheet.cell(row=i, column=9).value, sheet.cell(row=i, column=12).value, sheet.cell(row=i, column=13).value))

                self.label.setText('File successfully loaded')
                self.done_dialog()
                courses = self.groupabbr(courses)
                self.coursesconnector = courses
            
        except FileNotFoundError:
            self.label.setText('Problems with the input file')
            info(
                'Problems with the input file')
        except InvalidFileException:
            self.label.setText('Unsupported file format')  

    def on_gen_clicked(self):
        if self.finallist:
            outlist = [p for p in product(*self.finallist) if not any(one & two for one, two in combinations(p, 2))]
            self.label.setText('Generating schedule...')
            temp = 1
            with open('result.txt', 'w') as file:
                for k in outlist:
                    file.write('\n' + 'Schedule #' + str(temp) + '\n')
                    file.write('-------------------' + '\n')
                    for l in k:
                        file.write(str(l) + '\n')
                    temp = temp + 1
            self.label.setText('Results successfully saved as result.txt')
            # TODO: workaround
            #sleep(1)
            #self.label.setText('Current list is ' + str([y[0].abbr for y in self.finallist]))
        else:
            self.label.setText('Cannot create a schedule with given courses')

    def checkxl(self, sheet):
        """ A simple check to compare the input file's properties to the "standard's" """
        
        return sheet.cell(row=1, column=1).value == 'Course Abbr' and sheet.max_column == 13

    def groupabbr(self, inlist):
        
        """ Forms a list of course lists based on the course abbreviation and section (s/t) """
        
        result = defaultdict(lambda: defaultdict(list))

        for obj in inlist:
            result[obj.abbr][reg.search(obj.st).group(1)].append(obj)
        
        inlist = [(list(r.values())[0][0].abbr, list(r.values())) for r in result.values()]
        info('Courses were sorted')

        return inlist

    def initui(self):
        
        # Defining labels
        # -------------------------------------
        self.label = QLabel(self)
        self.label.setAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        if not self.finallist:
            self.label.setText('File is not loaded')

        # Defining buttons
        # -------------------------------------
        info('Defining buttons')
        open_button = QPushButton("Open xlsx and load it")
        add_button = QPushButton("Edit needed courses")
        help_button = QPushButton("Help")
        about_button = QPushButton("About")
        gen_button = QPushButton("Generate!")
        gen_button.setStyleSheet(
            'QPushButton {background-color: #c7f439; color: red; font-size: 50px;}')

        # Dealing with the interface of the app
        # -------------------------------------
        grid = QGridLayout() 
        grid.addWidget(open_button)
        grid.addWidget(add_button)
        grid.addWidget(help_button)
        grid.addWidget(about_button)
        grid.addWidget(self.label)
        grid.addWidget(gen_button)
        self.setLayout(grid)

        # Slot calls
        # ------------------------------------
        help_button.clicked.connect(self.on_help_clicked)
        add_button.clicked.connect(self.on_add_clicked)
        about_button.clicked.connect(self.on_about_clicked)
        open_button.clicked.connect(self.on_open_clicked)
        gen_button.clicked.connect(self.on_gen_clicked)

        # Setting window properties
        # ------------------------------------
        self.setGeometry(600, 300, 300, 270)
        self.center()
        self.setWindowTitle('nu-schedule')
        self.setWindowIcon(QIcon('res/logo.ico'))
        self.show()
        info('Main window loaded')


def main():
    info('Starting the app')
    app = QApplication(argv)
    widget = UI()
    return app.exec_()

if __name__ == '__main__':
    exit(main())
